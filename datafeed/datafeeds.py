from datetime import datetime
from openpyxl import load_workbook
import my_dropbox 
import my_helpers 
import requests
import re 
nullToZero = my_helpers.nullToZero
getDropbox = my_dropbox.getDropbox

class DataFeed(object):
    type = ""

    def __init__(self):
        self.state = {}
    def log(self, message):
        print(self.type + ": " + str(datetime.now()) + " " + message) 
    def getData(self):
        pass
    def populateDB(self, conn):
        pass

class PlayerFeed(DataFeed):
    PLAYER_SQL = """Insert Into Player(Id, firstName, lastname, currentteamid, jersey, isactive, position, heightfeet, heightinches, weightpounds, dateofbirth, nbadebutyear,yearspro, college )
        Values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) On Conflict (id) Do Update 
        Set firstName=%s,
            lastName=%s,
            currentteamid=%s,
            jersey=%s,
            isactive=%s,
            position=%s,
            heightfeet=%s,
            heightinches=%s,
            weightpounds=%s,
            dateofbirth=%s,
            nbadebutyear=%s,
            yearspro=%s,
            college=%s"""
     
    PLAYER_URL = "http://data.nba.net/10s/prod/v1/2018/players.json"
    
    def __init__(self):
        super().__init__()
        self.type = "PlayerFeed"
        
    def getData(self):
        response = requests.get(self.PLAYER_URL) 
        data = response.json()
        self.state['data'] = data   
    
    def populateDB(self, conn):
        cur = conn.cursor()
        records = self.state['data']
        for data in records['league']['standard']:
            cur.execute(self.PLAYER_SQL, (data['personId'], data['firstName'], data['lastName'],  my_helpers.nullToZero(data['teamId']), data['jersey'], data['isActive'], data['pos'], data['heightFeet'], data['heightInches'], data['weightPounds'], data['dateOfBirthUTC'], data['nbaDebutYear'], data['yearsPro'], data['collegeName'], data['firstName'], data['lastName'],  my_helpers.nullToZero(data['teamId']), data['jersey'], data['isActive'], data['pos'], data['heightFeet'], data['heightInches'], data['weightPounds'], data['dateOfBirthUTC'], data['nbaDebutYear'], data['yearsPro'], data['collegeName']))
        cur.close()    
        
class TeamFeed(DataFeed):
    TEAM_URL = "http://data.nba.net/10s/prod/v1/2018/teams.json"
    TEAM_SQL = """Insert Into Team(Id, City, AltCityName, Tricode, Urlname, nickname, Conference, Division, fullName)
        Values(%s, %s, %s, %s, %s, %s, %s, %s, %s) On Conflict Do Nothing"""
    
    def __init__(self):
        super().__init__()
        self.type = "TeamFeed"
    
    def getData(self):
        response = requests.get(self.TEAM_URL) 
        data = response.json()
        self.state['data'] = data
    
    def populateDB(self, conn):
        cur = conn.cursor()
        data = self.state['data']
        for record in my_helpers.filterListOfDicts(data['league']['standard'], 'isNBAFranchise', True):
            cur.execute(self.TEAM_SQL, (data['teamId'], data['city'], data['altCityName'],  data['tricode'], data['urlName'], data['nickname'], data['confName'], data['divName'], data['fullName']))
        cur.close()
              
class Game_TeamStatsPerGameFeed(DataFeed):
    GAMESQL_INSERT = """
    Insert Into game(id, gamedate, minutes, possesions, pace, url)
    Values(%s, %s, %s, %s, %s, %s) On Conflict Do Nothing;
    """
    GAMETEAMSTATSSQL_INSERT = """
    Insert Into gameteamstats (gameid, team, q1, q2, q3, q4, ot1, ot2, ot3, ot4, ot5,
    fga, fg, threesa, threes, fta, ft, ofr, dr, rebounds, assists, personalfouls, turnovers, blocks, oeff, deff, rest, points, venue)
    Values(%s,%s, %s, %s, %s, %s, %s,%s,%s,%s, %s,%s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s,
     %s, %s, %s, %s, %s, %s) On Conflict Do Nothing;
    """
    UPDATE_IMPLIED_DATA = """
    Select update_implied_data_values();
    """
    def __init__(self):
        super().__init__()
        self.type = "Game_TeamStatsPerGameFeed"
    
    def getData(self):
        workbook = self.__getWorkbook()
        worksheet = self.__getWorksheet(workbook)
        colHeaders = self.__getColHeaders(worksheet)
        rowCount = worksheet.max_row
        colCount = (len(colHeaders)) 
        
        self.state['sheet'] = worksheet
        self.state['headers'] = colHeaders
        self.state['rowCount'] = rowCount
        self.state['colCount'] = colCount  
    
    def populateDB(self, conn):
        sheet = self.state['sheet'] 
        columnHeaders = self.state['headers']
        rowCount = self.state['rowCount']
        columnCount = self.state['colCount'] 
        
        cur = conn.cursor()
        rowData = {}
        rowNum = 0
        for row in sheet.rows:
            rowNum += 1
            if rowNum < 2:
                continue # skip the first row, which is the headers
            colNum = 0    
            for cell in row:
                rowData[columnHeaders[colNum]] = cell.value
                colNum+=1
            cur.execute(self.GAMESQL_INSERT, (rowData['GAME-ID'],rowData['DATE'],rowData['MIN'],(rowData['POSS'] * 2),rowData['PACE'],rowData['BOXSCOREURL']))
            cur.execute(self.GAMETEAMSTATSSQL_INSERT, (rowData['GAME-ID'],rowData['TEAM'],rowData['1Q'],rowData['2Q'],rowData['3Q'],rowData['4Q'],nullToZero(rowData['OT1']),nullToZero(rowData['OT2']),nullToZero(rowData['OT3']),nullToZero(rowData['OT4']),nullToZero(rowData['OT5']),rowData['FGA'],rowData['FG'],rowData['3PA'],rowData['3P'],rowData['FTA'],rowData['FT'],rowData['OR'],rowData['DR'],rowData['TOT'],rowData['A'],rowData['PF'],rowData['TOTO'],rowData['BL'],rowData['OEFF'],rowData['DEFF'],rowData['TEAMRESTDAYS'],rowData['F'],rowData['VENUE']))
        cur.execute(self.UPDATE_IMPLIED_DATA) #db_side funtion to derive the implied data from one table and populate the the others """
        cur.close()         
    
    def __getWorkbook(self):
        dropbxPath = "/NBA/18-19-nba-team/"
        staticFileNameElem = "-nba-season-team-feed.xlsx"
        fileName = my_helpers.getMonthDayYear(-1) + staticFileNameElem
        fullFileName = dropbxPath + fileName
        localPath = "./data/" + fileName
        dropbx = getDropbox()
        dropbx.files_download_to_file(localPath,fullFileName)     
        workbook = load_workbook(filename="./data/" + my_helpers.getMonthDayYear(-1) + "-nba-season-team-feed.xlsx", read_only=True)
        return workbook
    
    def __getWorksheet(self, wb):
        sheet = wb.active
        return sheet
    
    def __getColHeaders(self, sheet):
        columnCount = sheet.max_column
        columnHeaders = ["null"]
        starter=1
        for i in range(1, columnCount + 1):
            cell = sheet.cell(row=1, column=i)
            value = cell.value
            if value == 'STARTING LINEUPS' or value == None:
                value = 'STARTER' + str(starter)
            starter +=1
            value = re.sub(r"[\n\t\s]*", "", value)
            columnHeaders.append(value)
        return columnHeaders[1:]
  
class PlayerStatsPerGameFeed(DataFeed):
    PLAYERSTATSSQL_INSERT = """
    Insert Into gameplayerstats(gameid, playerid, playerfullname, ownteam,opposingteam, venue, starter, minutes, fga,
							    fg,threesa, threes, ft, fta, ofr, dfr, rebounds, assists, pf,
							    steals, blocks, turnovers, points, usage, rest)
                Values(%s,%s, %s, %s, %s, %s, %s,%s,%s,%s, %s,%s, %s, %s, %s, %s, %s,
                       %s, %s,%s, %s, %s, %s, %s, %s) On Conflict Do Nothing;
    """
    def __init__(self):
        super().__init__()
        self.type = "PlayerStatsPerGameFeed"
    
    def getData(self):
        workbook = self.__getWorkbook()
        worksheet = self.__getWorksheet(workbook)
        colHeaders = self.__getColHeaders(worksheet)
        rowCount = worksheet.max_row
        colCount = (len(colHeaders)) 
        
        self.state['sheet'] = worksheet
        self.state['headers'] = colHeaders
        self.state['rowCount'] = rowCount
        self.state['colCount'] = colCount  
    
    def populateDB(self, conn):
        sheet = self.state['sheet'] 
        columnHeaders = self.state['headers']
        rowCount = self.state['rowCount']
        columnCount = self.state['colCount'] 
        
        cur = conn.cursor()
        rowData = {}
        rowNum = 0
        for row in sheet.rows:
            rowNum += 1
            if rowNum < 2:
                continue # skip the first row, which is the headers
            colNum = 0    
            for cell in row:
                rowData[columnHeaders[colNum]] = cell.value
                colNum+=1
            cur.execute(self.PLAYERSTATSSQL_INSERT, (rowData['GAME-ID'],rowData['PLAYER-ID'], rowData['PLAYERFULLNAME'],rowData['OWNTEAM'],rowData['OPPONENTTEAM'],
                                                     rowData['VENUE(R/H)'],rowData['STARTER(Y/N)'],rowData['MIN'],rowData['FGA'],
                                                     rowData['FG'],rowData['3PA'],rowData['3P'],
                                                     rowData['FT'],rowData['FTA'],rowData['OR'],rowData['DR'],rowData['TOT'],
                                                     rowData['A'],rowData['PF'],rowData['ST'],rowData['BL'],rowData['TO'],rowData['PTS'],
                                                     rowData['USAGERATE(%)'],rowData['DAYSREST']))
        cur.close()     
    """for row in range(2, rowCount + 1):
            for col in range(1, columnCount + 1):
                rowData[columnHeaders[col]] = sheet.cell(row=row, column=col).value

              """
        
    
    def __getWorkbook(self):
        dropbxPath = "/NBA/18-19-nba-player/"
        staticFileNameElem = "-nba-season-player-feed.xlsx"
        fileName = my_helpers.getMonthDayYear(-1) + staticFileNameElem
        fullFileName = dropbxPath + fileName
        localPath = "./data/" + fileName
        dropbx = getDropbox()
        dropbx.files_download_to_file(localPath,fullFileName)     
        workbook = load_workbook(filename="./data/" + my_helpers.getMonthDayYear(-1) + "-nba-season-player-feed.xlsx", read_only=True)
        return workbook
    
    def __getWorksheet(self, wb):
        sheet = wb.active
        return sheet
    
    def __getColHeaders(self, sheet):
        columnCount = sheet.max_column
        columnHeaders = ["null"]
        starter=1
        for i in range(1, columnCount + 1):
            cell = sheet.cell(row=1, column=i)
            value = cell.value
            value = re.sub(r"[\n\t\s]*", "", value)
            columnHeaders.append(value)
        return columnHeaders[1:]
  
if __name__ == "__main__":
    print("compiles")